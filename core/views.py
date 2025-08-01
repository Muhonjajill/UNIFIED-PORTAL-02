from django.shortcuts import render, get_list_or_404, redirect

from core.priority_rules import determine_priority
from .models import File, FileAccessLog
from django.http import FileResponse, JsonResponse, HttpResponse
from .forms import FileUploadForm, ProblemCategoryForm, TicketForm
from django.contrib.auth.decorators import login_required, permission_required, user_passes_test
from django.db.models import Count, Q
from django.utils.timezone import now
from django.db.models.functions import TruncMonth
from core.models import FileCategory
import os
from collections import Counter
from django.contrib.auth.models import User, Group
from django.utils.decorators import method_decorator
from .forms import UserUpdateForm, ProfileUpdateForm, TerminalForm,TerminalUploadForm, VersionControlForm, FileUploadForm, CustomUserCreationForm, LoginForm, OTPForm,TicketEditForm,TicketComment, TicketCommentForm, TicketForm
from django.views import View
import csv
from .models import Customer, Region, Terminal, Unit, SystemUser, Zone, ProblemCategory, VersionControl,VersionComment, Report, Ticket, Profile, EmailOTP,TicketComment
from django.core.mail import send_mail, EmailMultiAlternatives, EmailMessage
from django.utils.html import strip_tags
from django.contrib import messages
from datetime import datetime
from django.utils.dateparse import parse_date
from django.utils import timezone
import calendar
from django.shortcuts import get_object_or_404
from mimetypes import guess_type
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth import login as auth_login, authenticate
from django import forms
import random
from django.core.exceptions import PermissionDenied
from core.utils import can_user_access_file
from .utils import is_director  
import json
import pandas as pd
from email.mime.image import MIMEImage
from datetime import datetime, timedelta
from io import BytesIO
from core import models
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from django.views.decorators.http import require_POST


def in_group(user, group_name):
    return user.is_authenticated and (user.is_superuser or user.groups.filter(name=group_name).exists())
def is_director(user):
    return in_group(user, 'Director')
def is_manager(user):
    return in_group(user, 'Manager')
def is_staff(user):
    return in_group(user, 'Staff')

@user_passes_test(is_director)
def admin_dashboard(request):
    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'update_role':
            user_id = request.POST.get('user_id')
            new_role = request.POST.get('new_role')
            user = get_object_or_404(User, id=user_id)
            user.groups.clear()
            group, _ = Group.objects.get_or_create(name=new_role)
            user.groups.add(group)
            messages.success(request, f"{user.username}'s role updated to {new_role}.")

        elif action == 'delete_user':
            user_id = request.POST.get('user_id')
            user = get_object_or_404(User, id=user_id)
            user.delete()
            messages.success(request, f"User {user.username} has been deleted.")

    users = User.objects.exclude(id=request.user.id)

    context = {
        'users': users,
        'total_users': User.objects.count(),
        'total_files': File.objects.count(),
        'open_tickets': Ticket.objects.filter(status='open').count(),
    }
    return render(request, 'accounts/admin_dashboard.html', context)

@user_passes_test(is_director)
def manage_file_categories(request):
    categories = FileCategory.objects.all()

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'create':
            name = request.POST.get('name')
            icon = request.POST.get('icon')
            if name and icon:  
                FileCategory.objects.create(name=name, icon=icon)
                messages.success(request, f'Category "{name}" created successfully.')
                return redirect('manage_file_categories')

        elif action == 'update':
            category_id = request.POST.get('category_id')
            new_name = request.POST.get('new_name')
            new_icon = request.POST.get('icon')
            category = get_object_or_404(FileCategory, id=category_id)
            
            # Preserve the existing icon if no new icon is selected
            category.name = new_name
            if new_icon:
                category.icon = new_icon  
            category.save()
            messages.success(request, f'Category "{new_name}" updated successfully.')
            return redirect('manage_file_categories')

        elif action == 'delete':
            category_id = request.POST.get('category_id')
            category = get_object_or_404(FileCategory, id=category_id)
            category.delete()
            messages.success(request, f'Category "{category.name}" deleted.')
            return redirect('manage_file_categories')

    return render(request, 'accounts/manage_file_categories.html', {
        'categories': categories
    })


@user_passes_test(is_director)
def create_user(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        first_name = request.POST.get('first_name') 
        last_name = request.POST.get('last_name') 
        email = request.POST.get('email')
        password = request.POST.get('password')
        role = request.POST.get('role')

        if User.objects.filter(username=username).exists():
            messages.error(request, 'Username already exists.')
            return redirect('create_user')

        user = User.objects.create_user(
            username=username,
            email=email,
            password=password
        )

        user.first_name = first_name
        user.last_name = last_name
        user.save()

        group, _ = Group.objects.get_or_create(name=role)
        user.groups.add(group)

        messages.success(request, f"{role} user created successfully.")
        return redirect('admin_dashboard')

    return render(request, 'accounts/create_user.html')

class RegistrationForm(forms.ModelForm):
    password = forms.CharField(widget=forms.PasswordInput)
    class Meta:
        model = User
        fields = ['username', 'email', 'password']


def register_view(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('login')
    else:
        form = CustomUserCreationForm()
    return render(request, 'accounts/register.html', {'form': form})


def login_view(request):
    form = LoginForm()

    if request.method == 'POST':
        form = LoginForm(request.POST)
        
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data["password"]
            user = authenticate(username=username, password=password)
            if user:
                request.session['pre_otp_user'] = user.id 
                otp = str(random.randint(100000, 999999))
                EmailOTP.objects.update_or_create(user=user, defaults={'otp': otp, 'created_at': timezone.now()})
                
                # Prepare HTML and plain text content
                subject = 'Your OTP Code'
                html_content = f"""
                    <html>
                    <head>
                        <style>
                        @import url('https://fonts.googleapis.com/css?family=Rubik:400,700&display=swap');
                        body {{
                            font-family: 'Rubik', 'Helvetica Neue', Arial, sans-serif;
                            margin: 0;
                            padding: 0;
                            background: linear-gradient(120deg, #1e3c72 0%, #2a5298 100%);
                            min-height: 100vh;
                        }}
                        .email-container {{
                            max-width: 600px;
                            margin: 60px auto;
                            background: #fff;
                            border-radius: 16px;
                            box-shadow: 0 8px 32px 0 rgba(44,62,80,0.12);
                            overflow: hidden;
                            padding: 0 0 40px 0;
                            animation: fadeIn 1s;
                        }}
                        .header-bar {{
                            width: 100%;
                            height: 52px;
                            background: linear-gradient(90deg,#3498db 30%, #e74c3c 80%);
                            display: flex;
                            align-items: center;
                            justify-content: center;
                        }}
                        .logo {{
                            height: 100%;
                            margin: 16px auto 8px auto;
                            display: block;
                            filter: drop-shadow(0 2px 8px rgba(52,152,219,0.12));
                        }}
                        h2 {{
                            text-align: center;
                            color: #1e3c72;
                            font-size: 30px;
                            font-weight: 700;
                            margin: 16px 0 0 0;
                            letter-spacing: 1px;
                        }}
                        .accent-divider {{
                            width: 56px;
                            height: 4px;
                            background: linear-gradient(90deg, #3498db, #e74c3c);
                            border-radius: 2px;
                            margin: 18px auto 24px auto;
                        }}
                        p {{
                            color: #34495e;
                            font-size: 17px;
                            margin: 20px 0;
                            text-align: center;
                        }}
                        .otp-container {{
                            background: linear-gradient(96deg, #e74c3c 60%, #3498db);
                            margin: 35px auto 25px auto;
                            padding: 30px 20px;
                            border-radius: 12px;
                            max-width: 250px;
                            box-shadow: 0 6px 16px 0 rgba(231,76,60,0.08);
                            text-align: center;
                            border: 0.5px solid #f2f2f2;
                            position: relative;
                            animation: fadeIn 2s;
                        }}
                        .otp-glow {{
                            font-size: 36px;
                            font-weight: bold;
                            letter-spacing: 4px;
                            padding: 14px 34px;
                            background: #fff2f0;
                            color: #e74c3c;
                            border-radius: 10px;
                            margin: 20px 0 12px 0;
                            box-shadow: 0 0 25px 7px rgba(231,76,60,0.09);
                            position: relative;
                            animation: shimmer 2.5s linear infinite;
                        }}
                        @keyframes shimmer {{
                            0% {{ box-shadow: 0 0 20px 4px #fffCC2; }}
                            50% {{ box-shadow: 0 0 38px 7px #ffeabf; }}
                            100% {{ box-shadow: 0 0 20px 4px #fffCC2; }}
                        }}
                        .otp-expiry {{
                            color: #fff;
                            font-size: 15px;
                            margin-top: 18px;
                            font-style: italic;
                        }}
                        .cta-button {{
                            margin: 35px auto 0 auto;
                            display: block;
                            width: max-content;
                            background: linear-gradient(90deg, #3498db, #9b59b6);
                            color: #fff !important;
                            font-size: 20px;
                            text-decoration: none;
                            padding: 18px 38px;
                            border-radius: 8px;
                            font-weight: 700;
                            letter-spacing: 1px;
                            box-shadow: 0 4px 12px 0 rgba(41,128,185,0.14);
                            transition: background 0.25s, transform 0.2s;
                        }}
                        .cta-button:hover {{
                            background: #2a70b8;
                            transform: scale(1.06);
                        }}
                        .footer {{
                            margin-top: 60px;
                            font-size: 14px;
                            color: #98a4b3;
                            text-align: center;
                            padding: 32px 14px 0 14px;
                        }}
                        .footer strong {{ color: #34495e; }}
                        .footer a {{
                            color: #2980b9;
                            text-decoration: none;
                            transition: color 0.25s;
                        }}
                        .footer a:hover {{ color: #e74c3c; }}
                        @keyframes fadeIn {{
                            from {{ opacity: 0;transform: translateY(40px); }}
                            to {{ opacity: 1;transform: translateY(0);  }}
                        }}
                        @media only screen and (max-width: 600px) {{
                            .email-container {{ padding: 0 0 20px 0; border-radius: 0; margin: 0; }}
                            .otp-glow {{ font-size: 26px; padding: 10px 10px; }}
                            .otp-container {{ padding: 15px 5px; }}
                        }}
                        </style>
                    </head>
                    <body>
                        <div class="email-container">
                        <div class="header-bar"></div>
                        <img src="cid:logo" alt="BRITS Logo" class="logo" />
                        <h2>Hi {user.username},</h2>
                        <div class="accent-divider"></div>
                        <p>
                            We received a login request for your account.<br>
                            Please use the One-Time Password (OTP) below to complete your login.
                        </p>
                        <div class="otp-container">
                            <div class="otp-glow">{otp}</div>
                            <div class="otp-expiry">Your code expires in <strong>5 minutes</strong>.</div>
                        </div>
                        <p>
                            For your security, do not share this OTP code with anyone.<br>
                            If you did not make this request, please <a href="https://yourapp.com/security">secure your account</a>.
                        </p>
                        <a href="https://yourapp.com/security" class="cta-button">
                            Review Account Activity
                        </a>
                        <div class="footer">
                            Best regards,<br>
                            <strong>Blue River Technology Solutions</strong><br>
                            <span style="font-size:12px;">✉️ This inbox is not monitored for replies</span>
                        </div>
                        </div>
                    </body>
                    </html>
                """
                text_content = strip_tags(html_content)

                # Create the email object
                email = EmailMultiAlternatives(
                    subject,
                    text_content,
                    'no-reply@yourapp.com',
                    [user.email],
                )
                email.attach_alternative(html_content, "text/html")

                # Attach the logo as an inline image using MIMEImage
                with open('static/icons/logo.png', 'rb') as logo_file:
                    logo_data = logo_file.read()
                    logo = MIMEImage(logo_data, name='logo.png')
                    logo.add_header('Content-ID', '<logo>')  

                    email.attach(logo) 

                email.send()
                return JsonResponse({'status': 'otp_sent'})
            else:
                return JsonResponse({'status': 'error', 'message': 'Invalid username or password'})
        else:
            return JsonResponse({'status': 'error', 'message': 'Invalid form input'})
    return render(request, 'accounts/login.html', {'form': form})


def verify_otp_view(request):
    user_id = request.session.get('pre_otp_user')
    if not user_id:
        return JsonResponse({'status': 'error', 'message': 'Session expired. Please login again.'})

    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'User not found'})

    if request.method == "POST":
        form = OTPForm(request.POST)
        if form.is_valid():
            otp_input = form.cleaned_data["otp"]
            otp_instance = EmailOTP.objects.filter(user=user).first()

            if otp_instance:
                if otp_input != otp_instance.otp:
                    return JsonResponse({'status': 'error', 'message': 'Invalid OTP'})
                elif otp_instance.is_expired():
                    return JsonResponse({'status': 'error', 'message': 'Expired OTP'})
                else:
                    auth_login(request, user)
                    if 'pre_otp_user' in request.session:
                        del request.session['pre_otp_user']
                    otp_instance.delete()  
                    return JsonResponse({'status': 'verified', 'redirect_url': '/pre_dashboards/'})
            else:
                return JsonResponse({'status': 'error', 'message': 'OTP not found'})

        return JsonResponse({'status': 'error', 'message': 'Invalid OTP input'})
    
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})


@login_required
def edit_file(request, file_id):
    #if not request.user.has_perm('core.change_file'):
        #raise PermissionDenied("You do not have permission to edit this file")
    if file.access_level == 'confidential' and file.uploaded_by != request.user and not request.user.is_superuser:
        raise PermissionDenied("This confidential file can only be edited by its uploader.")
    file = get_object_or_404(File, pk=file_id)

    if request.method == 'POST':
        form = FileUploadForm(request.POST, request.FILES, instance=file)
        if form.is_valid():
            form.save()
            return redirect('view_files')
    else:
        form = FileUploadForm(instance=file)

    return render(request, 'edit_file.html', {'form': form, 'file': file})

def pre_dashboards(request):
    return render(request, 'core/pre_dashboards.html')


#@user_passes_test(is_viewer)
def user_list_view(request):
    users = User.objects.all()
    paginator = Paginator(users, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'core/file_management/user_list.html', {'page_obj': page_obj})

@user_passes_test(is_staff)
def user_detail(request, user_id):
    user = get_object_or_404(User, id=user_id)
    return render(request, 'core/file_management/user_detail.html', {'user': user})

@user_passes_test(is_staff)
def edit_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        user.username = request.POST['username']
        user.email = request.POST['email']
        user.is_active = 'is_active' in request.POST
        user.save()
        messages.success(request, 'User updated successfully!')
        return redirect('user_list')
    return render(request, 'core/file_management/edit_user.html', {'user': user})

@permission_required('auth.delete_user', raise_exception=True)
def delete_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    user.delete()
    messages.success(request, 'User deleted successfully!')
    return redirect('user_list')

@login_required
def file_management_dashboard(request):
    # Get all files that are not deleted
    files = File.objects.filter(is_deleted=False)
    
    # Count file extensions
    ext_counter = Counter()
    for f in files:
        ext = os.path.splitext(f.file.name)[1].lower()  # Get file extension
        ext_counter[ext] += 1

    # File types data with counts
    file_types = [
        {"type": "PDF Documents", "ext": ".pdf", "icon": "pdf", "count": ext_counter.get(".pdf", 0)},
        {"type": "Word Documents", "ext": ".docx", "icon": "docx", "count": ext_counter.get(".docx", 0)},
        {"type": "Images", "ext": ".jpg", "icon": "image", "count": ext_counter.get(".jpg", 0) + ext_counter.get(".png", 0)},
        {"type": "Excel Sheets", "ext": ".xlsx", "icon": "xlsx", "count": ext_counter.get(".xlsx", 0)},
        {"type": "Others", "ext": "other", "icon": "file", "count": sum(ext_counter.values()) - (
            ext_counter.get(".pdf", 0) + ext_counter.get(".docx", 0) + ext_counter.get(".jpg", 0) +
            ext_counter.get(".png", 0) + ext_counter.get(".xlsx", 0)
        )},
    ]

    # Count files by category (excluding deleted ones)
    categories = FileCategory.objects.annotate(
        file_count=Count('file', filter=Q(file__is_deleted=False))
    )    

    user = request.user
    # Get recent files, ordered by upload date
    recent_files = File.objects.filter(is_deleted=False).order_by('-upload_date')[:5]
    
    # Check visibility based on access level
    visible_files = []
    for file in recent_files:
        if file.access_level == 'public' or user.is_superuser:
            file.extension = os.path.splitext(file.file.name)[1] 
            visible_files.append(file)
        elif file.access_level == 'restricted' and file.authorized_users.filter(id=user.id).exists() or user.is_superuser:
            file.extension = os.path.splitext(file.file.name)[1]
            visible_files.append(file)
        elif file.access_level == 'confidential' and (file.uploaded_by == user or user.is_superuser):
            file.extension = os.path.splitext(file.file.name)[1]
            visible_files.append(file)

    # Return the render with visible files
    return render(request, 'core/file_management/dashboard.html', {
        'categories': categories,
        'recent_files': visible_files,  
        'file_types': file_types,
        'user_name': request.user.username  
    })

@login_required
def file_list_view(request, category_name=None):
    user = request.user
    files = File.objects.filter(is_deleted=False)

    visible_files = []

    # Iterate over each file to check access rights
    for file in files:
        if file.can_user_access(user):
            visible_files.append(file)

    # Filter by category if provided
    if category_name:
        visible_files = [file for file in visible_files if file.category.name.lower() == category_name.lower()]

    # Sort files
    sort_option = request.GET.get('sort')
    if sort_option == 'recent':
        visible_files.sort(key=lambda f: f.upload_date, reverse=True)
    else:
        visible_files.sort(key=lambda f: f.title)

    # Paginate files
    paginator = Paginator(visible_files, 10)
    page = request.GET.get('page')
    try:
        paginated_files = paginator.page(page)
    except PageNotAnInteger:
        paginated_files = paginator.page(1)
    except EmptyPage:
        paginated_files = paginator.page(paginator.num_pages)

    # Fetch categories for the filter dropdown
    categories = FileCategory.objects.all()

    return render(request, 'core/file_management/file_list.html', {
        'files': paginated_files,
        'categories': categories,
        'active_category': category_name,
    })


def search(request):
    query = request.GET.get('q', '')
    files = File.objects.filter(title__icontains=query, is_deleted=False)
    categories = FileCategory.objects.filter(name__icontains=query)
    users = User.objects.filter(username__icontains=query)

    context = {
        'query': query,
        'files': files,
        'categories': categories,
        'users': users,
    }
    return render(request, 'core/file_management/search_result.html', context)


@login_required
def preview_file(request, file_id):
    file = get_object_or_404(File, id=file_id, is_deleted=False)

    # Access control
    if file.access_level == 'public':
        pass  
    elif file.access_level == 'restricted':
        if not file.can_user_access(request.user):
            raise PermissionDenied("Access denied.")
    elif file.access_level == 'confidential':
        if request.user != file.uploaded_by and not request.user.is_superuser:
            raise PermissionDenied("Access denied for confidential file.")
    else:
        raise PermissionDenied("Unknown access level.")

    # Log access regardless of type
    FileAccessLog.objects.create(file=file, accessed_by=request.user)

    # Serve file if it's previewable
    mime_type, _ = guess_type(file.file.name)
    if mime_type in ['application/pdf', 'image/jpeg', 'image/png', 'image/gif']:
        return FileResponse(file.file.open('rb'), content_type=mime_type)

    # Unsupported type
    return render(request, 'core/file_management/unsupported_preview.html', {'file': file})

@login_required
def delete_file(request, file_id):
    if not request.user.has_perm('core.delete_file'):
        raise PermissionDenied('You do not have permission to delete this file')
    file = get_object_or_404(File, id=file_id, is_deleted=False)

    if request.method == "POST":
        file.is_deleted = True
        file.save()
        messages.success(request, "File deleted successfully.")
        return redirect('file_list')
    
    return redirect('file_list')
    
@login_required
@permission_required('core.add_file', raise_exception=True)
def upload_file_view(request):
    if request.method == 'POST':
        form = FileUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file_instance = form.save(commit=False)
            file_instance.uploaded_by = request.user
            file_instance.save()
            return redirect('file_list')
    else:
        form = FileUploadForm()
    
    return render(request, 'core/file_management/upload_file.html', {'form': form})

#@user_passes_test(is_staff)
def profile_view(request):
    context = {
        'user': request.user,
        'user_form': UserUpdateForm(instance=request.user),
        'profile_form': ProfileUpdateForm(instance=request.user.profile),
    }

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'accounts/profile_content.html', context)

    return render(request, 'accounts/profile.html', context)


@method_decorator(login_required, name='dispatch')
class SettingsView(View):
    def get(self, request):
        user_form = UserUpdateForm(instance=request.user)
        profile, created = Profile.objects.get_or_create(user=request.user)
        profile_form = ProfileUpdateForm(instance=profile)
        return render(request, 'accounts/settings.html', {
            'user_form': user_form,
            'profile_form': profile_form
        })

    def post(self, request):
        user_form = UserUpdateForm(request.POST, instance=request.user)
        profile, created = Profile.objects.get_or_create(user=request.user)
        profile_form = ProfileUpdateForm(request.POST, request.FILES, instance=profile)
        if user_form.is_valid() and profile_form.is_valid():
            user_form.save()
            profile_form.save()
            return redirect('profile_view')
        return render(request, 'accounts/settings.html', {
            'user_form': user_form,
            'profile_form': profile_form
        })


# Ticketing Views
def ticketing_dashboard(request):
    now = timezone.localtime(timezone.now())
    print(f"Now (timezone aware): {now}")

    # Define period starts
    day_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    week_start = day_start - timedelta(days=day_start.weekday())
    month_start = day_start.replace(day=1)
    year_start = day_start.replace(month=1, day=1)
  

    # Time-based ticket counts
    time_data = {
        'day': Ticket.objects.filter(created_at__gte=day_start).count(),
        'week': Ticket.objects.filter(created_at__gte=week_start).count(),
        'month': Ticket.objects.filter(created_at__gte=month_start).count(),
        'year': Ticket.objects.filter(created_at__gte=year_start).count(),
    }

    # Aggregated counts
    status_counts = Ticket.objects.values('status').annotate(count=Count('id'))
    priority_counts = Ticket.objects.values('priority').annotate(count=Count('id'))

    # Monthly ticket creation trends
    monthly_trends = (
        Ticket.objects
        .annotate(month=TruncMonth('created_at'))
        .values('month')
        .annotate(count=Count('id'))
        .order_by('month')
    )

    # Top terminals with the most tickets
    terminal_data = (
        Ticket.objects
        .values('terminal__branch_name')
        .annotate(count=Count('id'))
        .order_by('-count')[:10]
    )

    region_data = (
        Ticket.objects
        .values('terminal__region__name')  
        .annotate(count=Count('id'))
        .order_by('-count')
    )

    # Top categories
    category_data = (
        Ticket.objects
        .values('problem_category__name')
        .annotate(count=Count('id'))
        .order_by('-count')[:10]
    )

    # Top customers
    customer_data = (
        Ticket.objects
        .values('terminal__customer__name')
        .annotate(count=Count('id'))
        .order_by('-count')[:10]
    )

    # Overview Data: Aggregated ticket counts for different time periods
    overview_data = [
        {'label': 'Daily', 'count': time_data['day']},
        {'label': 'Weekly', 'count': time_data['week']},
        {'label': 'Monthly', 'count': time_data['month']},
        {'label': 'Yearly', 'count': time_data['year']},
    ]
    category_time_data = [
        {'category': d['problem_category__name'], 'daily_count': d['count']} 
        for d in category_data  
    ]
    kpi_data = [
        ("Daily", "dailyCount", "fa-sun"),
        ("Weekly", "weeklyCount", "fa-calendar-week"),
        ("Monthly", "monthlyCount", "fa-calendar-alt"),
        ("Yearly", "yearlyCount", "fa-calendar"),
    ]
    # Prepare data for JSON serialization
    context = {
        "kpi_data": kpi_data,
        'status_data': json.dumps(list(status_counts)),
        'priority_data': json.dumps(list(priority_counts)),
        'monthly_data': json.dumps([{'month': calendar.month_abbr[d['month'].month], 'count': d['count']} for d in monthly_trends if d['month']]),
        'terminal_data': json.dumps([{'terminal': d['terminal__branch_name'], 'count': d['count']} for d in terminal_data]),
        'region_data': json.dumps([{'region': d['terminal__region__name'], 'count': d['count']} for d in region_data]), 
        'time_data': json.dumps(time_data),
        'category_data': json.dumps([{'category': d['problem_category__name'], 'count': d['count']} for d in category_data]),
        'customer_data': json.dumps([{'customer': d['terminal__customer__name'], 'count': d['count']} for d in customer_data]),
        'overview_data': json.dumps(overview_data),  
        'category_time_data': json.dumps(category_time_data), 
    }

    return render(request, 'core/helpdesk/ticketing_dashboard.html', context)

def statistics_view(request):
    today = timezone.now()
    print(f"timezone.now() in view: {today} (aware: {timezone.is_aware(today)})")
   
    # Get the selected filter values from the request
    time_period = request.GET.get('time-period', "all_time")
    customer_filter = request.GET.get('customer', 'all')
    terminal_filter = request.GET.get('terminal', 'all')
    region_filter = request.GET.get('region', 'all')
    
    # Determine the date range for filtering based on time period
    if time_period == 'today':
        start_date = today.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = today.replace(hour=23, minute=59, second=59, microsecond=999999)
    elif time_period == 'yesterday':
        start_date = (today - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = (today - timedelta(days=1)).replace(hour=23, minute=59, second=59, microsecond=999999)
    elif time_period == 'lastweek':
        start_date = today - timedelta(days=today.weekday() + 7)
        start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = start_date + timedelta(days=6, hours=23, minutes=59, seconds=59, microseconds=999999)
    elif time_period == 'lastmonth':
        start_date = (today.replace(day=1) - timedelta(days=1)).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        end_date = today.replace(day=1) - timedelta(microseconds=1) # End of last month
    elif time_period == 'lastyear':
        start_date = today.replace(year=today.year - 1, month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        end_date = today.replace(year=today.year - 1, month=12, day=31, hour=23, minute=59, second=59, microsecond=999999)
    elif time_period == 'all_time':
        # All time: no date filtering
        start_date = None
        end_date = None
    else:
        start_date = today.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = today.replace(hour=23, minute=59, second=59, microsecond=999999)

    # Query data with applied filters
    tickets = Ticket.objects.all()

    if customer_filter != 'all':
        tickets = tickets.filter(terminal__customer__id=customer_filter)
    if terminal_filter != 'all':
        tickets = tickets.filter(terminal__id=terminal_filter)
    if region_filter != 'all':
        tickets = tickets.filter(terminal__region__id=region_filter)
    
    # Apply the date range filter only if not "all_time"
    if time_period != 'all_time':
        tickets = tickets.filter(created_at__range=[start_date, end_date])

    # Calculate ticket statuses (you can customize these status names)
    ticket_statuses = tickets.values('status').annotate(status_count=Count('status'))
    status_labels = [status['status'] for status in ticket_statuses]
    status_counts = [status['status_count'] for status in ticket_statuses]
    
    # Data aggregation for different timeframes
    days = [today - timedelta(days=i) for i in range(7)]  # Last 7 days
    tickets_per_day = [tickets.filter(created_at__date=day.date()).count() for day in days]
    
    weekdays = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
    tickets_per_weekday = [tickets.filter(created_at__week_day=i).count() for i in range(1, 8)] 
    
    hours = [f"{i}-{i+1}" for i in range(24)]
    tickets_per_hour = [tickets.filter(created_at__hour=i).count() for i in range(24)]
    
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    tickets_per_month = [tickets.filter(created_at__month=i+1).count() for i in range(12)]
   
    # Collect distinct years from the filtered tickets, then sort
    years = sorted(list(set(ticket.created_at.year for ticket in tickets.iterator()))) 
    tickets_per_year = [tickets.filter(created_at__year=year).count() for year in years]

    tickets_per_terminal = tickets.values('terminal__branch_name').annotate(ticket_count=Count('id'))
    print("Tickets Per Terminal:", tickets_per_terminal)  

    ticket_categories = tickets.values('problem_category').annotate(ticket_count=Count('id'))
    
    # Fetch terminals, customers, and regions
    terminals_qs = Terminal.objects.select_related('customer', 'region').all()
    terminals = [
        {
            'id': terminal.id,
            'branch_name': terminal.branch_name,
            'customer__name': terminal.customer.name if terminal.customer else 'N/A',
            'region__name': terminal.region.name if terminal.region else 'N/A',
        }
        for terminal in terminals_qs
    ]

    customers = list(Customer.objects.values('id', 'name'))
    regions = list(Region.objects.values('id', 'name'))


    # Pack data for the frontend
    data = {
        'ticketsPerTerminal': [
            {
                'branch_name': entry['terminal__branch_name'],
                'count': entry['ticket_count']
            }
            for entry in tickets_per_terminal
        ],
        'ticketCategories': {
            'labels': [entry['problem_category'] for entry in ticket_categories],
            'data': [entry['ticket_count'] for entry in ticket_categories]
        },
        'ticketStatuses': {
            'labels': status_labels,
            'data': status_counts,
        },
        'days': [day.strftime('%Y-%m-%d') for day in days],
        'ticketsPerDay': tickets_per_day,
        'weekdays': weekdays,
        'ticketsPerWeekday': tickets_per_weekday,
        'hours': hours,
        'ticketsPerHour': tickets_per_hour,
        'months': months,
        'ticketsPerMonth': tickets_per_month,
        'years': years,
        'ticketsPerYear': tickets_per_year,
        'terminals': terminals,
        'customers': customers,
        'regions': regions,
    }
    
    # Check if the request is AJAX
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse(data, safe=False)
    
    return render(request, 'core/helpdesk/statistics.html', {
        'customers': customers,
        'terminals': terminals,
        'regions': regions,
        "time_period": time_period,
        'selected_customer': customer_filter,
        'selected_terminal': terminal_filter,
        'selected_region': region_filter,
        'data_json': json.dumps(data, ensure_ascii=False),
    })

def export_report(request):
    # Get the selected filter values from the request
    time_period = request.GET.get('time-period', 'today')
    customer_filter = request.GET.get('customer', 'all')
    terminal_filter = request.GET.get('terminal', 'all')
    region_filter = request.GET.get('region', 'all')

    # Get the current time
    today = timezone.now()

    # Determine the date range for filtering based on time period
    if time_period == 'today':
        start_date = today.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = today.replace(hour=23, minute=59, second=59, microsecond=999999)
    elif time_period == 'yesterday':
        start_date = (today - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = (today - timedelta(days=1)).replace(hour=23, minute=59, second=59, microsecond=999999)
    elif time_period == 'lastweek':
        start_date = today - timedelta(days=today.weekday() + 7)
        start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = start_date + timedelta(days=6, hours=23, minutes=59, seconds=59, microseconds=999999)
    elif time_period == 'lastmonth':
        start_date = (today.replace(day=1) - timedelta(days=1)).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        end_date = today.replace(day=1) - timedelta(microseconds=1) 
    elif time_period == 'lastyear':
        start_date = today.replace(year=today.year - 1, month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        end_date = today.replace(year=today.year - 1, month=12, day=31, hour=23, minute=59, second=59, microsecond=999999)
    else:
        start_date = today.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = today.replace(hour=23, minute=59, second=59, microsecond=999999)

    # Query tickets with the applied filters
    tickets = Ticket.objects.all()

    # Filter by time range
    tickets = tickets.filter(created_at__range=[start_date, end_date])

    # Filter by customer if selected
    if customer_filter != 'all':
        tickets = tickets.filter(terminal__customer__id=customer_filter)

    # Filter by terminal if selected
    if terminal_filter != 'all':
        tickets = tickets.filter(terminal__id=terminal_filter)

    # Filter by region if selected
    if region_filter != 'all':
        tickets = tickets.filter(terminal__region__id=region_filter)

    # Calculate ticket statistics for per day, per weekday, etc.
    tickets_per_day = [tickets.filter(created_at__date=today - timedelta(days=i)).count() for i in range(7)]
    tickets_per_weekday = [tickets.filter(created_at__week_day=i).count() for i in range(1, 8)]  
    tickets_per_hour = [tickets.filter(created_at__hour=i).count() for i in range(24)]
    tickets_per_month = [tickets.filter(created_at__month=i+1).count() for i in range(12)]
    years = sorted(list(set(ticket.created_at.year for ticket in tickets.iterator())))
    tickets_per_year = [tickets.filter(created_at__year=year).count() for year in years]

    # Create the Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ticket Statistics"

    # Add headers to the sheet
    headers = ['Ticket ID', 'Customer', 'Terminal', 'Region', 'Created At']
    ws.append(headers)

    # Add ticket data rows
    for ticket in tickets:
        # Correct access to customer via terminal
        customer_name = ticket.terminal.customer.name if ticket.terminal and ticket.terminal.customer else 'No Customer'

        # Correct access to region via terminal
        region_name = ticket.terminal.region.name if ticket.terminal and ticket.terminal.region else 'No Region'

        # Ensure created_at is timezone naive
        created_at_naive = ticket.created_at.replace(tzinfo=None)

        ws.append([ticket.id, customer_name, ticket.terminal.cdm_name, region_name, created_at_naive])

    # Add the breakdown of tickets (Per Day, Per Week, Per Month, etc.)
    ws.append([])
    ws.append(['Statistics Breakdown'])
    ws.append(['Tickets per Day (Last 7 Days)'])
    for i, count in enumerate(tickets_per_day):
        ws.append([f"Day {7-i}", count])

    ws.append(['Tickets per Weekday'])
    weekdays = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
    for i, count in enumerate(tickets_per_weekday):
        ws.append([weekdays[i], count])

    ws.append(['Tickets per Hour'])
    for i, count in enumerate(tickets_per_hour):
        ws.append([f"{i}-{i+1}", count])

    ws.append(['Tickets per Month'])
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    for i, count in enumerate(tickets_per_month):
        ws.append([months[i], count])

    ws.append(['Tickets per Year'])
    for i, count in enumerate(tickets_per_year):
        ws.append([years[i], count])

    # Create a response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="ticket_statistics.xlsx"'
    wb.save(response)
    return response


def tickets(request):
    query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '') 
    
    tickets = Ticket.objects.select_related('problem_category').filter(
        Q(title__icontains=query) |
        Q(description__icontains=query) |
        Q(problem_category__name__icontains=query)
    ).order_by('-created_at')

    if status_filter:
        tickets = tickets.filter(status=status_filter)

    #pagination
    paginator = Paginator(tickets, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)  

    return render(request, 'core/helpdesk/tickets.html', {
        'tickets': page_obj,
        'search_query': query,
        'status_filter': status_filter
    })



def create_ticket(request):
    if request.method == 'POST':
        form = TicketForm(request.POST)
        if form.is_valid():
            print(form.cleaned_data)
            ticket = form.save(commit=False)

            # Handle FK safely
            category_name = ticket.problem_category.name if ticket.problem_category else 'other'

            # Auto-determine priority
            ticket.priority = determine_priority(
                ticket.issue_type,
                category_name,
                ticket.description
            )

            ticket.created_by = request.user

            # Auto-populate related fields
            if ticket.terminal:
                ticket.customer = ticket.terminal.customer
                ticket.region = ticket.terminal.region

            ticket.save()
            return redirect('create_ticket' if 'create_another' in request.POST else 'tickets')
    else:
        # Handle pre-populated terminal
        terminal_id = request.GET.get('terminal_id')  
        if terminal_id:
            form = TicketForm(terminal_id=terminal_id)
        else:
            form = TicketForm()

    return render(request, 'core/helpdesk/create_ticket.html', {'form': form})

def get_terminal_details(request, terminal_id):
    try:
        terminal = Terminal.objects.get(id=terminal_id)
        response_data = {
            'customer_id': terminal.customer.id if terminal.customer else None,
            'region_id': terminal.region.id if terminal.region else None,
        }
        return JsonResponse(response_data)
    except Terminal.DoesNotExist:
        return JsonResponse({'error': 'Terminal not found'}, status=404)
    
def ticket_detail(request, ticket_id):
    ticket = get_object_or_404(Ticket, id=ticket_id)
    comments = ticket.comments.order_by('-created_at')
    form = TicketEditForm(instance=ticket)
    comment_form = TicketCommentForm()

    if request.method == 'POST':
        if 'add_comment' in request.POST:
            comment_form = TicketCommentForm(request.POST)
            if comment_form.is_valid():
                comment = comment_form.save(commit=False)
                comment.ticket = ticket
                comment.created_by = request.user
                comment.save()
                return redirect('ticket_detail', ticket_id=ticket.id)

        elif 'edit_ticket' in request.POST:
            form = TicketEditForm(request.POST, instance=ticket)
            if form.is_valid():
                form.save()
                return redirect('ticket_detail', ticket_id=ticket.id)
            else:
                print("EDIT FORM ERRORS:", form.errors)  # 👈 debug

    elif is_director(request.user) or is_manager(request.user):
        form = TicketForm(instance=ticket)

    context = {
        'ticket': ticket,
        'form': form,
        'comments': comments,
        'comment_form': comment_form,
        'is_admin': request.user.is_superuser,
        'is_editor': request.user.groups.filter(name='Editor').exists(),
        'can_resolve': request.user.groups.filter(name='Resolver').exists(),
    }

    if is_director(request.user) or is_manager(request.user):
        return render(request, 'core/helpdesk/ticket_detail.html', context)

    elif is_staff(request.user):
        if ticket.status in ['resolved', 'closed']:
            return render(request, 'core/helpdesk/ticket_detail.html', context)
        else:
            return render(request, 'core/helpdesk/permission_denied.html')

    return render(request, 'core/helpdesk/permission_denied.html')


@login_required
def edit_comment(request, comment_id):
    comment = get_object_or_404(TicketComment, id=comment_id)

    if request.user != comment.created_by and not request.user.is_superuser:
        messages.error(request, "You don't have permission to edit this comment.")
        return redirect('ticket_detail', ticket_id=comment.ticket.id)

    if request.method == 'POST':
        form = TicketCommentForm(request.POST, instance=comment)
        if form.is_valid():
            form.save()
            messages.success(request, "Comment updated successfully.")
            return redirect('ticket_detail', ticket_id=comment.ticket.id)
    else:
        form = TicketCommentForm(instance=comment)

    return render(request, 'core/helpdesk/edit_comment.html', {'form': form, 'comment': comment})


@login_required
def delete_comment(request, comment_id):
    comment = get_object_or_404(TicketComment, id=comment_id)

    if request.user != comment.created_by and not request.user.is_superuser:
        messages.error(request, "You don't have permission to delete this comment.")
        return redirect('ticket_detail', ticket_id=comment.ticket.id)

    if request.method == 'POST':
        ticket_id = comment.ticket.id
        comment.delete()
        messages.success(request, "Comment deleted.")
        return redirect('ticket_detail', ticket_id=ticket_id)



@login_required
def resolve_ticket_view(request, ticket_id):
    ticket = get_object_or_404(Ticket, id=ticket_id)
    resolution = request.POST.get('resolution', '').strip()

    # Check if the user is authorized to resolve the ticket
    if is_director(request.user) or is_manager(request.user):
        # Admins and Editors can resolve tickets
        if ticket.status != 'resolved':
            ticket.resolution = resolution
            ticket.status = 'closed'
            ticket.resolved_by = request.user  
            ticket.resolved_at = timezone.now()
            ticket.save()
            messages.success(request, 'Ticket resolved successfully!')
            return redirect('ticket_detail', ticket_id=ticket.id)
        else:
            messages.error(request, 'Ticket already resolved')
            return render(request, 'core/helpdesk/error.html')

    elif request.user.has_perm('can_resolve_ticket'):
        # Custom permission check
        if ticket.status != 'resolved':
            ticket.status = 'resolved'
            ticket.resolved_by = request.user  
            ticket.resolved_at = timezone.now() 
            ticket.save()
            messages.success(request, 'Ticket resolved successfully!')
            return redirect('ticket_detail', ticket_id=ticket.id)
        else:
            messages.error(request, 'Ticket already resolved!')
            return render(request, 'core/helpdesk/error.html')

    # If the user doesn't have permission
    messages.error(request, 'You do not have permission to resolve this ticket.')
    return render(request, 'core/helpdesk/permission_denied.html')


@user_passes_test(is_director)
def delete_ticket(request, ticket_id):
    ticket = get_object_or_404(Ticket, id=ticket_id)
    ticket.delete()
    messages.success(request, "Ticket deleted successfully.")
    return redirect('tickets')

def ticket_statuses(request):
    return render(request, 'core/helpdesk/ticket_statuses.html')


def tickets_by_status(request, status):
    tickets = Ticket.objects.filter(status__iexact=status.replace('-', '_'))
    return render(request, 'core/helpdesk/ticket_by_status.html', {
        'status': status.title().replace('-', ' '),
        'tickets': tickets
    })

def problem_category(request):
    print(">>> problem_category view reached")
    query = request.GET.get('search', '')
    categories = ProblemCategory.objects.filter(name__icontains=query)
    print(f"Categories found: {categories.count()}")
    
    # Pagination setup
    paginator = Paginator(categories, 10)  
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'core/helpdesk/problem_category.html', {
        'page_obj': page_obj, 
        'search_query': query,
    })

@user_passes_test(is_director)
def create_problem_category(request):
    if request.method == 'POST':
        print("POST received:", request.POST) 
        form = ProblemCategoryForm(request.POST)
        if form.is_valid():
            category = form.save()
            print("Category saved!")

            # Redirect based on which button was clicked
            if 'create_another' in request.POST:
                return redirect('create_problem_category')
            return redirect('problem_category')  
        else:
            print("Form errors:", form.errors)
    else:
        form = ProblemCategoryForm()

    return render(request, 'core/helpdesk/create_problem_category.html', {'form': form})

@user_passes_test(is_director)
def edit_problem_category(request, category_id):
    category = get_object_or_404(ProblemCategory, pk=category_id)
    if request.method == 'POST':
        form = ProblemCategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            return redirect('problem_category')
    else:
        form = ProblemCategoryForm(instance=category)

    return render(request, 'core/helpdesk/edit_problem_category.html', {'form': form})

@user_passes_test(is_director)
def delete_problem_category(request, category_id):
    category = get_object_or_404(ProblemCategory, id=category_id)
    category.delete()
    messages.success(request, "Problem category deleted successfully.")
    return redirect('problem_category')

# Master Data Views
def customers(request):
    if request.method == "POST" and request.FILES.get("file"):
        csv_file = request.FILES["file"]
        decoded_file = csv_file.read().decode("utf-8").splitlines()
        reader = csv.DictReader(decoded_file)

        for row in reader:
            name = row.get("name", "").strip()
            if name: 
                Customer.objects.create(name=name)

        messages.success(request, "Customers uploaded successfully!")

    # Pagination setup
    all_customers = Customer.objects.exclude(name__exact="").exclude(name__isnull=True).order_by('id')
    paginator = Paginator(all_customers, 10)  
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, "core/helpdesk/customers.html", {"customers": page_obj})

@user_passes_test(is_director)
def create_customer(request):
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        if name:
            Customer.objects.create(name=name)
            messages.success(request, "Customer added successfully.")
            return redirect("customers")
        else:
            messages.error(request, "Customer name is required.")

    return render(request, "core/helpdesk/create_customer.html")

@user_passes_test(is_director)
def delete_customer(request, id):
    customer = get_object_or_404(Customer, id=id)
    customer.delete()
    messages.success(request, "Customer deleted successfully.")
    return redirect('customers')

def regions(request):
    if request.method == 'POST':
        name = request.POST.get('region_name')
        if name:
            Region.objects.create(name=name)
            return redirect('regions')

    # Fetch all regions
    all_regions = Region.objects.all().order_by('id')

    # Pagination setup
    paginator = Paginator(all_regions, 10)  
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'core/helpdesk/regions.html', {'regions': page_obj})

@user_passes_test(is_director)
def delete_region(request, region_id):
    region = get_object_or_404(Region, id=region_id)
    region.delete()
    messages.success(request, "Region deleted successfully.")
    return redirect('regions')

def terminals(request):
    form = TerminalForm()
    upload_form = TerminalUploadForm()

    # Get all the required objects to pass to the template
    customers = Customer.objects.all()
    regions = Region.objects.all()
    zones = Zone.objects.all()

    # Handle POST request for terminal creation or CSV upload
    if request.method == 'POST':
        # Handle terminal creation
        if 'create' in request.POST or 'create_another' in request.POST:
            print("Form submitted")
            form = TerminalForm(request.POST)
            if form.is_valid():
                print("Form is valid")
                try:
                    form.save()
                    messages.success(request, "Terminal created successfully.")

                    # Handle the "Create & Add Another" functionality
                    if 'create_another' in request.POST:
                        return redirect('terminals')

                    return redirect('terminals')  # Redirect to the terminal list after creation
                except Exception as e:
                    messages.error(request, f"Error creating terminal: {e}")
                    print(f"Error creating terminal: {e}")
            else:
                print("Form is not valid")
                print("Form errors:", form.errors) 
        # Handle CSV upload for terminals
        elif 'upload_file' in request.POST:
            upload_form = TerminalUploadForm(request.POST, request.FILES)
            if upload_form.is_valid():
                file = upload_form.cleaned_data['file']
                try:
                    if file.name.endswith('.csv'):
                        df = pd.read_csv(file)
                    else:
                        df = pd.read_excel(file)

                    # Process each row and create terminal objects
                    for _, row in df.iterrows():
                        Terminal.objects.create(
                            customer=Customer.objects.get(name=row['customer']),
                            branch_name=row['branch_name'],
                            cdm_name=row['cdm_name'],
                            serial_number=row['serial_number'],
                            region=Region.objects.get(name=row['region']),
                            model=row['model'],
                            zone=Zone.objects.get(name=row['zone']),
                        )
                    messages.success(request, "Terminals imported successfully.")
                except Exception as e:
                    messages.error(request, f"Error importing file: {e}")
                return redirect('terminals')

    # GET request: Display the page with all terminals
    all_terminals = Terminal.objects.all().order_by('id') 
    paginator = Paginator(all_terminals, 10)  
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Pass the required objects to the template
    return render(request, 'core/helpdesk/terminals.html', {
        'form': form,
        'upload_form': upload_form,
        'terminals': page_obj,
        'customers': customers,
        'regions': regions,
        'zones': zones
    })


@require_POST
def edit_terminal(request, terminal_id):
    terminal = get_object_or_404(Terminal, id=terminal_id)
    terminal.customer_id = request.POST.get('customer')
    terminal.branch_name = request.POST.get('branch_name')
    terminal.cdm_name = request.POST.get('cdm_name')
    terminal.serial_number = request.POST.get('serial_number')
    terminal.region_id = request.POST.get('region')
    terminal.model = request.POST.get('model')
    terminal.zone_id = request.POST.get('zone')
    
    try:
        terminal.save()
        messages.success(request, "Terminal updated successfully.")
    except Exception as e:
        messages.error(request, f"Error updating terminal: {e}")
    
    return redirect('terminals')

def fetch_tickets(request, terminal_id):
    tickets = Ticket.objects.filter(terminal_id=terminal_id).values('id', 'title')
    
    if tickets:
        return JsonResponse({"success": True, "tickets": list(tickets)})
    else:
        return JsonResponse({"success": False, "message": "No tickets found."})


@user_passes_test(is_director)
def delete_terminal(request, terminal_id):
    terminal = get_object_or_404(Terminal, id=terminal_id)
    terminal.delete()
    messages.success(request, "Terminal removed successfully.")
    return redirect('terminals')

def units(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description')
        if name and description:
            Unit.objects.create(name=name, description=description)
        return redirect('units')

    all_units = Unit.objects.all().order_by('id')
    
    paginator = Paginator(all_units, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'core/helpdesk/units.html', {'page_obj': page_obj})

@user_passes_test(is_director)
def delete_unit(request, unit_id):
    unit = get_object_or_404(Unit, id=unit_id)
    unit.delete()
    messages.success(request, "Unit removed successfully.")
    return redirect('units')

def system_users(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        role = request.POST.get('role')
        users = User.objects.all()
        if username and email and role:
            SystemUser.objects.create(username=username, email=email, role=role)
        return redirect('system_users')

    all_users = User.objects.all().order_by('id')
    # Add pagination: Show 10 users per page
    paginator = Paginator(all_users, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'core/helpdesk/users.html', {'page_obj': page_obj})

@user_passes_test(is_director)
def delete_system_user(request, user_id):
    user = get_object_or_404(User, id=user_id)
    if request.user == user:
        messages.error(request, "You cannot delete your own account.")
    else:
        user.delete()
        messages.success(request, "User deleted successfully.")
    return redirect('system_users')

def zones(request):
    if request.method == 'POST':
        name = request.POST.get('name')

        if name: 
            Zone.objects.create(name=name)
            messages.success(request, "Zone created successfully.")
            return redirect('zones')
        else:
            messages.error(request, "Name is required.")

    all_zones = Zone.objects.all().order_by('id')

    # Add pagination: Show 10 zones per page
    paginator = Paginator(all_zones, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'core/helpdesk/zones.html', {
        'page_obj': page_obj,
    })

@user_passes_test(is_director)
def delete_zone(request, zone_id):
    zone = get_object_or_404(Zone, id=zone_id)
    zone.delete()
    messages.success(request, "Zone deleted successfully.")
    return redirect('zones') 

def reports(request):

    tickets = Ticket.objects.prefetch_related('comments').all()

    customer = request.GET.get('customer')

    terminal_name = request.GET.get("terminal_name")

    region = request.GET.get('region')

    category = request.GET.get('category')

    filter_by_customer = False
    filter_by_terminal = False


    if customer and customer != 'All' and customer !="None":
        tickets = tickets.filter(customer_id=customer)

        filter_by_customer = True  
    if terminal_name: 
        tickets = tickets.filter(terminal__branch_name__icontains=terminal_name)
        filter_by_terminal = True  

    if region and region != 'All' and region != "None":
        tickets = tickets.filter(region_id=region)
    if category and category != 'All' and category !="None":
        tickets = tickets.filter(problem_category_id=category)

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if start_date:

        tickets = tickets.filter(created_at__date__gte=parse_date(start_date))

        tickets = tickets.filter(created_at__date__lte=parse_date(end_date))

      
    if request.GET.get('download') == 'excel':

        customer_name = Customer.objects.get(id=customer).name if customer and customer not in ['All', 'None'] else None
        terminal_filter = terminal_name if terminal_name else None

        return export_tickets_to_excel(
            tickets,
            include_terminal=filter_by_customer,
            customer_name=customer_name,
            terminal_name=terminal_filter,
            start_date=start_date,
            end_date=end_date
        )

    paginator = Paginator(tickets, 10)  # or however many per page
    page = request.GET.get('page')

    try:
        tickets_page = paginator.page(page)
    except PageNotAnInteger:
        tickets_page = paginator.page(1)
    except EmptyPage:
        tickets_page = paginator.page(paginator.num_pages)

    context = {
        'tickets': tickets_page,  
        'page_obj': tickets_page,  
        'customers': Customer.objects.all(),
        'terminals': Terminal.objects.all(),
        'regions': Region.objects.all(),
        'categories': ProblemCategory.objects.all(),
        'filter_by_customer': filter_by_customer,
        'filter_by_terminal': filter_by_terminal,
    }

    return render(request, 'core/helpdesk/reports.html', context)


def export_tickets_to_excel(tickets, include_terminal=False, customer_name=None, terminal_name=None, start_date=None, end_date=None):
    import openpyxl
    from openpyxl.utils import get_column_letter


    workbook = openpyxl.Workbook()

    sheet = workbook.active

    sheet.title = 'Tickets'

    # Headers
    headers = []
    if include_terminal:
        headers.append('Terminal')

    headers += ['Created At', 'Updated At', 'Problem Category', 'Status', 'Responsible', 'Description', 'Comments','Resolution']

    sheet.append(headers)


    # Data rows

    for ticket in tickets:
        comments_text = "\n".join([
            f"{comment.created_by.username if comment.created_by else 'Unknown'} ({comment.created_at.strftime('%Y-%m-%d')}): {comment.content}"
            for comment in ticket.comments.all()
        ])
        row = []
        if include_terminal:
            row.append(ticket.terminal.branch_name)
        row += [
            ticket.created_at.strftime('%Y-%m-%d %H:%M'),
            ticket.updated_at.strftime('%Y-%m-%d %H:%M'),
            str(ticket.problem_category),
            ticket.status,
            str(ticket.responsible),
            ticket.description,
            comments_text,
            ticket.resolution or "",
        ]
        sheet.append(row)

    comment_col_index = len(headers)  
    for row in sheet.iter_rows(min_row=2, min_col=comment_col_index, max_col=comment_col_index):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)
            
    # Dynamic filename
    name_part = "report"
    if customer_name:
        name_part = f"{customer_name.replace(' ', '_')}_report"
    elif terminal_name:
        name_part = f"{terminal_name.replace(' ', '_')}_report"

    date_part = ''
    if start_date and end_date:
        date_part = f"{start_date}_to_{end_date}"
    elif start_date:
        date_part = f"from_{start_date}"
    elif end_date:
        date_part = f"to_{end_date}"

    filename = f"{name_part}_{date_part or timezone.now().strftime('%Y-%m-%d')}.xlsx"


    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    response['Content-Disposition'] = f'attachment; filename={filename}'

    workbook.save(response)

    return response

def version_controls(request):
    form = VersionControlForm()

    if request.method == 'POST':
        if 'create' in request.POST or 'create_another' in request.POST:
            form = VersionControlForm(request.POST)
            if form.is_valid():
                form.save()
                if 'create_another' in request.POST:
                    form = VersionControlForm()
                else:
                    return redirect('version_controls')

    # Initial unfiltered queryset
    versions = VersionControl.objects.all().order_by('-created_at')

    # Handle AJAX filter request
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        terminal = request.GET.get('terminal')
        firmware = request.GET.get('firmware')
        app_version = request.GET.get('app_version')

        if terminal and terminal != 'All':
            versions = versions.filter(terminal=terminal)
        if firmware and firmware != 'All':
            versions = versions.filter(firmware=firmware)
        if app_version and app_version != 'All':
            versions = versions.filter(app_version=app_version)

        return render(request, 'core/helpdesk/partials/version_table.html', {
            'versions': versions
        })

    paginator = Paginator(versions, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Get terminals with their branch_name and id only
    terminals = VersionControl.objects.select_related('terminal').values('terminal__branch_name', 'terminal__id').distinct()

    context = {
        'form': form,
        'versions': versions,
        'page_obj':page_obj,
        'terminals': terminals,
        'firmwares': VersionControl.objects.values_list('firmware', flat=True).distinct(),
        'app_versions': VersionControl.objects.values_list('app_version', flat=True).distinct(),
    }
    return render(request, 'core/helpdesk/version_control.html', context)

def version_detail(request, pk):
    version = get_object_or_404(VersionControl, pk=pk)
    comments = version.comments.all().order_by('-created')  # Latest first

    if request.method == 'POST':
        comment_text = request.POST.get('comment')
        if comment_text:
            VersionComment.objects.create(version=version, text=comment_text)
        return redirect('version_detail', pk=pk)

    return render(request, 'core/helpdesk/version_detail.html', {
        'version': version,
        'comments': comments,
        
    })


def edit_version(request, pk):
    version = get_object_or_404(VersionControl, pk=pk)
    if request.method == 'POST':
        form = VersionControlForm(request.POST, instance=version)
        if form.is_valid():
            form.save()
            return redirect('version_detail', pk=pk)
    else:
        form = VersionControlForm(instance=version)

    return render(request, 'core/helpdesk/edit_version.html', {'form': form, 'version': version})

@user_passes_test(is_director)
def delete_version(request, pk):
    version = get_object_or_404(VersionControl, pk=pk)
    version.delete()
    return redirect('version_controls') 